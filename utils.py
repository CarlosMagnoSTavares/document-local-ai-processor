import os
import uuid
import pytesseract
from PIL import Image
import PyPDF2
from docx import Document as DocxDocument
import openpyxl
import httpx
import json
from pathlib import Path
from loguru import logger
from dotenv import load_dotenv

load_dotenv()

# Configuration
OLLAMA_BASE_URL = os.getenv("OLLAMA_BASE_URL", "http://localhost:11434")
UPLOAD_DIR = "uploads"
TEMP_DIR = "temp"
ALLOWED_EXTENSIONS = os.getenv("ALLOWED_EXTENSIONS", "pdf,jpg,jpeg,png,docx,xlsx,xls,doc").split(",")
MAX_FILE_SIZE = int(os.getenv("MAX_FILE_SIZE", "50")) * 1024 * 1024  # Convert MB to bytes

# Ensure directories exist
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(TEMP_DIR, exist_ok=True)

def is_allowed_file(filename: str) -> bool:
    """Check if file extension is allowed"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in [ext.strip().lower() for ext in ALLOWED_EXTENSIONS]

def save_uploaded_file(file_content: bytes, filename: str) -> str:
    """Save uploaded file and return the path"""
    file_extension = filename.rsplit('.', 1)[1].lower()
    unique_filename = f"{uuid.uuid4()}_{filename}"
    file_path = os.path.join(UPLOAD_DIR, unique_filename)
    
    with open(file_path, "wb") as f:
        f.write(file_content)
    
    return file_path

def extract_text_from_image(image_path: str) -> str:
    """Extract text from image using OCR"""
    try:
        logger.info(f"🖼️ VERBOSE: Starting OCR extraction from image: {image_path}")
        logger.info(f"🖼️ VERBOSE: File exists: {os.path.exists(image_path)}")
        
        if not os.path.exists(image_path):
            raise FileNotFoundError(f"Image file not found: {image_path}")
        
        # Verificar tamanho do arquivo
        file_size = os.path.getsize(image_path)
        logger.info(f"🖼️ VERBOSE: Image file size: {file_size} bytes")
        
        image = Image.open(image_path)
        logger.info(f"🖼️ VERBOSE: Image opened successfully - size: {image.size}, mode: {image.mode}")
        
        text = pytesseract.image_to_string(image, lang='por+eng')
        logger.info(f"🖼️ VERBOSE: OCR completed - extracted {len(text)} characters")
        logger.info(f"🖼️ VERBOSE: OCR preview: {text[:100]}..." if len(text) > 100 else f"🖼️ VERBOSE: OCR result: {text}")
        
        result = text.strip()
        logger.info(f"✅ VERBOSE: Image extraction successful - final length: {len(result)}")
        return result
    except Exception as e:
        logger.error(f"❌ VERBOSE: Error extracting text from image {image_path}: {e}")
        logger.error(f"❌ VERBOSE: Exception type: {type(e).__name__}")
        raise

def extract_text_from_pdf(pdf_path: str) -> str:
    """Extract text from PDF file"""
    try:
        logger.info(f"📄 VERBOSE: Starting PDF extraction from: {pdf_path}")
        logger.info(f"📄 VERBOSE: File exists: {os.path.exists(pdf_path)}")
        
        if not os.path.exists(pdf_path):
            raise FileNotFoundError(f"PDF file not found: {pdf_path}")
        
        # Verificar tamanho do arquivo
        file_size = os.path.getsize(pdf_path)
        logger.info(f"📄 VERBOSE: PDF file size: {file_size} bytes")
        
        text = ""
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            num_pages = len(pdf_reader.pages)
            logger.info(f"📄 VERBOSE: PDF has {num_pages} pages")
            
            for i, page in enumerate(pdf_reader.pages):
                page_text = page.extract_text()
                text += page_text + "\n"
                logger.info(f"📄 VERBOSE: Page {i+1} extracted {len(page_text)} characters")
        
        logger.info(f"📄 VERBOSE: PDF extraction completed - total {len(text)} characters")
        logger.info(f"📄 VERBOSE: PDF preview: {text[:100]}..." if len(text) > 100 else f"📄 VERBOSE: PDF result: {text}")
        
        result = text.strip()
        logger.info(f"✅ VERBOSE: PDF extraction successful - final length: {len(result)}")
        return result
    except Exception as e:
        logger.error(f"❌ VERBOSE: Error extracting text from PDF {pdf_path}: {e}")
        logger.error(f"❌ VERBOSE: Exception type: {type(e).__name__}")
        raise

def extract_text_from_docx(docx_path: str) -> str:
    """Extract text from DOCX file"""
    try:
        logger.info(f"📝 VERBOSE: Starting DOCX extraction from: {docx_path}")
        logger.info(f"📝 VERBOSE: File exists: {os.path.exists(docx_path)}")
        
        if not os.path.exists(docx_path):
            raise FileNotFoundError(f"DOCX file not found: {docx_path}")
        
        # Verificar tamanho do arquivo
        file_size = os.path.getsize(docx_path)
        logger.info(f"📝 VERBOSE: DOCX file size: {file_size} bytes")
        
        doc = DocxDocument(docx_path)
        text = ""
        paragraph_count = 0
        
        for paragraph in doc.paragraphs:
            paragraph_text = paragraph.text
            text += paragraph_text + "\n"
            if paragraph_text.strip():
                paragraph_count += 1
        
        logger.info(f"📝 VERBOSE: DOCX extraction completed - {paragraph_count} paragraphs, {len(text)} characters")
        logger.info(f"📝 VERBOSE: DOCX preview: {text[:100]}..." if len(text) > 100 else f"📝 VERBOSE: DOCX result: {text}")
        
        result = text.strip()
        logger.info(f"✅ VERBOSE: DOCX extraction successful - final length: {len(result)}")
        return result
    except Exception as e:
        logger.error(f"❌ VERBOSE: Error extracting text from DOCX {docx_path}: {e}")
        logger.error(f"❌ VERBOSE: Exception type: {type(e).__name__}")
        raise

def extract_text_from_excel(excel_path: str) -> str:
    """Extract text from Excel file"""
    try:
        logger.info(f"📊 VERBOSE: Starting Excel extraction from: {excel_path}")
        logger.info(f"📊 VERBOSE: File exists: {os.path.exists(excel_path)}")
        
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        
        # Verificar tamanho do arquivo
        file_size = os.path.getsize(excel_path)
        logger.info(f"📊 VERBOSE: Excel file size: {file_size} bytes")
        
        workbook = openpyxl.load_workbook(excel_path)
        text = ""
        total_rows = 0
        
        logger.info(f"📊 VERBOSE: Excel has {len(workbook.sheetnames)} sheets: {workbook.sheetnames}")
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"Sheet: {sheet_name}\n"
            sheet_rows = 0
            
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
                    sheet_rows += 1
            text += "\n"
            total_rows += sheet_rows
            logger.info(f"📊 VERBOSE: Sheet '{sheet_name}' has {sheet_rows} rows with data")
        
        logger.info(f"📊 VERBOSE: Excel extraction completed - {total_rows} total rows, {len(text)} characters")
        logger.info(f"📊 VERBOSE: Excel preview: {text[:100]}..." if len(text) > 100 else f"📊 VERBOSE: Excel result: {text}")
        
        result = text.strip()
        logger.info(f"✅ VERBOSE: Excel extraction successful - final length: {len(result)}")
        return result
    except Exception as e:
        logger.error(f"❌ VERBOSE: Error extracting text from Excel {excel_path}: {e}")
        logger.error(f"❌ VERBOSE: Exception type: {type(e).__name__}")
        raise

def extract_text_from_file(file_path: str, file_type: str) -> str:
    """Extract text from file based on its type"""
    file_type = file_type.lower()
    
    logger.info(f"🔍 VERBOSE: Starting text extraction from {file_type.upper()} file: {file_path}")
    
    if file_type in ['jpg', 'jpeg', 'png']:
        logger.info(f"🖼️ VERBOSE: Using OCR (Tesseract) for image processing")
        text = extract_text_from_image(file_path)
    elif file_type == 'pdf':
        logger.info(f"📄 VERBOSE: Using PyPDF2 for PDF processing")
        text = extract_text_from_pdf(file_path)
    elif file_type in ['docx', 'doc']:
        logger.info(f"📝 VERBOSE: Using python-docx for DOCX processing")
        text = extract_text_from_docx(file_path)
    elif file_type in ['xlsx', 'xls']:
        logger.info(f"📊 VERBOSE: Using openpyxl for Excel processing")
        text = extract_text_from_excel(file_path)
    else:
        logger.error(f"❌ VERBOSE: Unsupported file type: {file_type}")
        raise ValueError(f"Unsupported file type: {file_type}")
    
    logger.info(f"✅ VERBOSE: Text extraction completed. Extracted {len(text)} characters")
    logger.info(f"📄 VERBOSE: Text preview: {text[:300]}..." if len(text) > 300 else f"📄 VERBOSE: Full text: {text}")
    
    return text

async def send_prompt_to_ollama(prompt: str, context: str, model: str, format_response: str = None, example: str = None) -> tuple[str, str]:
    """Send prompt to Ollama and get response. Returns (llm_response, full_prompt)"""
    try:
        # Build enhanced prompt with strict formatting instructions
        format_instructions = ""
        if format_response and example:
            format_instructions = f"""

CRITICAL FORMATTING INSTRUCTIONS:
- You MUST respond ONLY with the exact JSON format specified below
- DO NOT include any explanations, introductions, or additional text
- DO NOT use markdown formatting or code blocks
- Respond with ONLY the JSON structure, nothing else
- Follow the exact pattern shown in the example

Required JSON Format: {format_response}
Example Response: {example}

Your response must be EXACTLY in this JSON format. No other text is allowed."""
        elif format_response:
            format_instructions = f"""

CRITICAL FORMATTING INSTRUCTIONS:
- You MUST respond ONLY with the exact JSON format specified below
- DO NOT include any explanations, introductions, or additional text
- DO NOT use markdown formatting or code blocks
- Respond with ONLY the JSON structure, nothing else

Required JSON Format: {format_response}

Your response must be EXACTLY in this JSON format. No other text is allowed."""

        full_prompt = f"""Context: {context}

Question: {prompt}{format_instructions}

Based on the context provided above, extract the required information and respond ONLY in the specified JSON format. Do not include any explanations or additional text."""
        
        logger.info(f"🤖 VERBOSE: Sending prompt to Ollama model '{model}'")
        logger.info(f"📄 VERBOSE: Context length: {len(context)} characters")
        logger.info(f"❓ VERBOSE: Prompt: {prompt}")
        if format_response:
            logger.info(f"📋 VERBOSE: Format required: {format_response}")
        if example:
            logger.info(f"💡 VERBOSE: Example provided: {example}")
        logger.debug(f"📝 VERBOSE: Full prompt: {full_prompt[:500]}..." if len(full_prompt) > 500 else f"📝 VERBOSE: Full prompt: {full_prompt}")
        
        async with httpx.AsyncClient(timeout=300) as client:
            logger.info(f"🔗 VERBOSE: Making request to {OLLAMA_BASE_URL}/api/generate")
            
            response = await client.post(
                f"{OLLAMA_BASE_URL}/api/generate",
                json={
                    "model": model,
                    "prompt": full_prompt,
                    "stream": False,
                    "options": {
                        "verbose": True,
                        "temperature": 0.1,  # Lower temperature for more consistent formatting
                        "top_p": 0.9,
                        "repeat_penalty": 1.1
                    }
                }
            )
            response.raise_for_status()
            result = response.json()
            
            llm_response = result.get("response", "").strip()
            logger.info(f"✅ VERBOSE: Ollama response received ({len(llm_response)} chars)")
            logger.info(f"💬 VERBOSE: Response preview: {llm_response[:200]}..." if len(llm_response) > 200 else f"💬 VERBOSE: Full response: {llm_response}")
            
            # Log additional Ollama metrics if available
            if "total_duration" in result:
                logger.info(f"⏱️ VERBOSE: Total duration: {result['total_duration']/1e9:.2f}s")
            if "load_duration" in result:
                logger.info(f"⚡ VERBOSE: Load duration: {result['load_duration']/1e9:.2f}s")
            if "prompt_eval_count" in result:
                logger.info(f"🔢 VERBOSE: Prompt tokens: {result['prompt_eval_count']}")
            if "eval_count" in result:
                logger.info(f"📊 VERBOSE: Response tokens: {result['eval_count']}")
            
            return llm_response, full_prompt
    except Exception as e:
        logger.error(f"❌ VERBOSE: Error sending prompt to Ollama: {e}")
        raise

def format_llm_response(llm_response: str, format_template: str, example: str = None) -> str:
    """Format LLM response according to specified format"""
    try:
        logger.info(f"🎨 VERBOSE: Starting response formatting")
        logger.info(f"📝 VERBOSE: Original LLM response: {llm_response}")
        logger.info(f"📋 VERBOSE: Expected format: {format_template}")
        
        # Clean the response
        cleaned_response = llm_response.strip()
        
        # Try to extract JSON from the response
        extracted_json = None
        
        # Method 1: Check if the entire response is valid JSON
        try:
            json.loads(cleaned_response)
            extracted_json = cleaned_response
            logger.info(f"✅ VERBOSE: Entire response is valid JSON")
        except json.JSONDecodeError:
            logger.info(f"🔍 VERBOSE: Response is not entirely JSON, attempting extraction")
            
        # Method 2: Try to find JSON within the response
        if extracted_json is None:
            import re
            
            # Look for JSON array patterns [...]
            if format_template.startswith('[') and format_template.endswith(']'):
                json_pattern = r'\[.*?\]'
                json_matches = re.findall(json_pattern, cleaned_response, re.DOTALL)
                
                for match in json_matches:
                    try:
                        json.loads(match)  # Validate JSON
                        extracted_json = match
                        logger.info(f"✅ VERBOSE: Found valid JSON array: {match}")
                        break
                    except json.JSONDecodeError:
                        continue
            
            # Look for JSON object patterns {...}
            elif format_template.startswith('{') and format_template.endswith('}'):
                json_pattern = r'\{.*?\}'
                json_matches = re.findall(json_pattern, cleaned_response, re.DOTALL)
                
                for match in json_matches:
                    try:
                        json.loads(match)  # Validate JSON
                        extracted_json = match
                        logger.info(f"✅ VERBOSE: Found valid JSON object: {match}")
                        break
                    except json.JSONDecodeError:
                        continue
        
        # Method 3: Try to extract specific values based on format template
        if extracted_json is None and format_template:
            logger.info(f"🔧 VERBOSE: Attempting value extraction based on template")
            try:
                # Parse the format template to understand expected structure
                template_obj = json.loads(format_template)
                
                if isinstance(template_obj, list) and len(template_obj) > 0 and isinstance(template_obj[0], dict):
                    # Handle array of objects
                    extracted_data = []
                    for key in template_obj[0].keys():
                        # Try to find values for this key in the response
                        value = extract_value_from_text(cleaned_response, key)
                        if value:
                            extracted_data.append({key: value})
                    
                    if extracted_data:
                        extracted_json = json.dumps(extracted_data, ensure_ascii=False)
                        logger.info(f"✅ VERBOSE: Extracted data based on template: {extracted_json}")
                
                elif isinstance(template_obj, dict):
                    # Handle single object
                    extracted_data = {}
                    for key in template_obj.keys():
                        value = extract_value_from_text(cleaned_response, key)
                        if value:
                            extracted_data[key] = value
                    
                    if extracted_data:
                        extracted_json = json.dumps(extracted_data, ensure_ascii=False)
                        logger.info(f"✅ VERBOSE: Extracted data based on template: {extracted_json}")
                        
            except json.JSONDecodeError:
                logger.warning(f"⚠️ VERBOSE: Could not parse format template as JSON")
        
        # Method 4: Fallback - try to construct response based on example
        if extracted_json is None and example:
            logger.info(f"🔧 VERBOSE: Using example as fallback guidance")
            try:
                example_obj = json.loads(example)
                if isinstance(example_obj, list) and len(example_obj) > 0 and isinstance(example_obj[0], dict):
                    extracted_data = []
                    for key in example_obj[0].keys():
                        value = extract_value_from_text(cleaned_response, key)
                        if value:
                            extracted_data.append({key: value})
                    
                    if extracted_data:
                        extracted_json = json.dumps(extracted_data, ensure_ascii=False)
                        logger.info(f"✅ VERBOSE: Constructed response based on example: {extracted_json}")
            except json.JSONDecodeError:
                logger.warning(f"⚠️ VERBOSE: Could not parse example as JSON")
        
        # Return the best result we found
        if extracted_json:
            logger.info(f"🎉 VERBOSE: Successfully formatted response: {extracted_json}")
            return extracted_json
        else:
            logger.warning(f"⚠️ VERBOSE: Could not extract valid JSON, returning original response")
            return cleaned_response
            
    except Exception as e:
        logger.error(f"❌ VERBOSE: Error formatting response: {e}")
        return llm_response

def extract_value_from_text(text: str, key: str) -> str:
    """Extract a value from text based on a key pattern"""
    import re
    
    # Common patterns to look for
    patterns = [
        # Direct patterns
        rf"{re.escape(key)}:\s*([^\n,}}]+)",
        rf"{re.escape(key)}\s*:\s*([^\n,}}]+)",
        rf"{re.escape(key)}\s*=\s*([^\n,}}]+)",
        
        # Date patterns (if key suggests it's a date)
        r"(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})",
        r"(\d{2,4}[/\-\.]\d{1,2}[/\-\.]\d{1,2})",
        
        # CNPJ patterns
        r"(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})",
        r"(\d{14})",
        
        # General number patterns
        r"(\d+[,\.]\d+)",
        r"(\d+)",
    ]
    
    # Try each pattern
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        if matches:
            # Clean and return the first match
            value = matches[0].strip()
            # Remove common trailing characters
            value = re.sub(r'[,;\.]+$', '', value)
            if value:
                return value
    
    return None

def cleanup_old_files():
    """Clean up old uploaded files and temporary files"""
    try:
        from datetime import datetime, timedelta
        import glob
        
        max_age_hours = int(os.getenv("MAX_RECORD_AGE_HOURS", "24"))
        cutoff_time = datetime.now() - timedelta(hours=max_age_hours)
        
        # Clean uploads directory
        for file_path in glob.glob(os.path.join(UPLOAD_DIR, "*")):
            if os.path.isfile(file_path):
                file_time = datetime.fromtimestamp(os.path.getctime(file_path))
                if file_time < cutoff_time:
                    os.remove(file_path)
                    logger.info(f"Cleaned up old file: {file_path}")
        
        # Clean temp directory
        for file_path in glob.glob(os.path.join(TEMP_DIR, "*")):
            if os.path.isfile(file_path):
                file_time = datetime.fromtimestamp(os.path.getctime(file_path))
                if file_time < cutoff_time:
                    os.remove(file_path)
                    logger.info(f"Cleaned up old temp file: {file_path}")
                    
    except Exception as e:
        logger.error(f"Error during cleanup: {e}")

def validate_file_size(file_size: int) -> bool:
    """Validate file size"""
    return file_size <= MAX_FILE_SIZE 

async def send_prompt_to_gemini(prompt: str, context: str, model: str, gemini_api_key: str, format_response: str = None, example: str = None) -> tuple[str, str]:
    """Send prompt to Google Gemini API and get response. Returns (llm_response, full_prompt)"""
    try:
        from google import genai
        
        logger.info(f"🤖 VERBOSE: Sending prompt to Google Gemini model '{model}'")
        logger.info(f"📄 VERBOSE: Context length: {len(context)} characters")
        logger.info(f"❓ VERBOSE: Prompt: {prompt}")
        if format_response:
            logger.info(f"📋 VERBOSE: Format required: {format_response}")
        if example:
            logger.info(f"💡 VERBOSE: Example provided: {example}")
        
        # Build enhanced prompt with strict formatting instructions
        format_instructions = ""
        if format_response and example:
            format_instructions = f"""

CRITICAL FORMATTING INSTRUCTIONS:
- You MUST respond ONLY with the exact JSON format specified below
- DO NOT include any explanations, introductions, or additional text
- DO NOT use markdown formatting or code blocks
- Respond with ONLY the JSON structure, nothing else
- Follow the exact pattern shown in the example

Required JSON Format: {format_response}
Example Response: {example}

Your response must be EXACTLY in this JSON format. No other text is allowed."""
        elif format_response:
            format_instructions = f"""

CRITICAL FORMATTING INSTRUCTIONS:
- You MUST respond ONLY with the exact JSON format specified below
- DO NOT include any explanations, introductions, or additional text
- DO NOT use markdown formatting or code blocks
- Respond with ONLY the JSON structure, nothing else

Required JSON Format: {format_response}

Your response must be EXACTLY in this JSON format. No other text is allowed."""

        full_prompt = f"""Context: {context}

Question: {prompt}{format_instructions}

Based on the context provided above, extract the required information and respond ONLY in the specified JSON format. Do not include any explanations or additional text."""
        
        # Create Gemini client
        client = genai.Client(api_key=gemini_api_key)
        
        logger.info(f"🔗 VERBOSE: Making request to Google Gemini API")
        logger.debug(f"📝 VERBOSE: Full prompt: {full_prompt[:500]}..." if len(full_prompt) > 500 else f"📝 VERBOSE: Full prompt: {full_prompt}")
        
        # Send request to Gemini
        response = client.models.generate_content(
            model=model,
            contents=full_prompt,
            config={
                'temperature': 0.1,  # Lower temperature for more consistent formatting
                'top_p': 0.9,
                'max_output_tokens': 2048,
            }
        )
        
        gemini_response = response.text.strip()
        logger.info(f"✅ VERBOSE: Gemini response received ({len(gemini_response)} chars)")
        logger.info(f"💬 VERBOSE: Response preview: {gemini_response[:200]}..." if len(gemini_response) > 200 else f"💬 VERBOSE: Full response: {gemini_response}")
        
        return gemini_response, full_prompt
        
    except Exception as e:
        logger.error(f"❌ VERBOSE: Error sending prompt to Gemini: {e}")
        raise 

async def list_gemini_models(gemini_api_key: str) -> dict:
    """List available Google Gemini models dynamically from API"""
    try:
        import httpx
        
        logger.info(f"🌟 VERBOSE: Fetching available Gemini models from API")
        
        # Call Google Gemini API to list models
        url = f"https://generativelanguage.googleapis.com/v1beta/models?key={gemini_api_key}"
        
        async with httpx.AsyncClient() as client:
            response = await client.get(url)
            
            if response.status_code == 200:
                data = response.json()
                models = []
                
                # Filter only generation models and extract relevant info
                for model in data.get('models', []):
                    model_name = model.get('name', '').replace('models/', '')
                    
                    # Only include generation models that support generateContent
                    supported_methods = model.get('supportedGenerationMethods', [])
                    if 'generateContent' in supported_methods:
                        # Extract description and create clean model info
                        description = model.get('description', '').split('.')[0]  # First sentence only
                        
                        models.append({
                            'name': model_name,
                            'description': description,
                            'version': model.get('version', 'latest'),
                            'input_token_limit': model.get('inputTokenLimit', 'unknown'),
                            'output_token_limit': model.get('outputTokenLimit', 'unknown')
                        })
                
                # Sort models by preference (newer versions first)
                models.sort(key=lambda x: (
                    '2.5' in x['name'],  # 2.5 first
                    '2.0' in x['name'],  # then 2.0
                    '1.5' in x['name'],  # then 1.5
                    'flash' in x['name']  # flash variants first
                ), reverse=True)
                
                logger.info(f"✅ VERBOSE: Successfully fetched {len(models)} Gemini models")
                return {
                    'status': 'success',
                    'models': models,
                    'total_models': len(models),
                    'recommended_model': models[0]['name'] if models else 'gemini-2.0-flash'
                }
            else:
                logger.error(f"❌ VERBOSE: Failed to fetch Gemini models. Status: {response.status_code}")
                logger.error(f"❌ VERBOSE: Response: {response.text}")
                return {
                    'status': 'error',
                    'message': f'Failed to fetch models from Gemini API: {response.status_code}',
                    'fallback_models': [
                        {
                            'name': 'gemini-2.0-flash',
                            'description': 'Latest multimodal model with next generation features',
                            'version': 'latest',
                            'size': None,
                            'modified': None,
                            'status': 'available'
                        },
                        {
                            'name': 'gemini-2.5-pro-preview', 
                            'description': 'Most powerful thinking model with enhanced reasoning',
                            'version': 'preview',
                            'size': None,
                            'modified': None,
                            'status': 'available'
                        },
                        {
                            'name': 'gemini-1.5-pro',
                            'description': 'Advanced model for complex reasoning tasks', 
                            'version': 'stable',
                            'size': None,
                            'modified': None,
                            'status': 'available'
                        },
                        {
                            'name': 'gemini-1.5-flash',
                            'description': 'Fast and versatile performance model',
                            'version': 'stable',
                            'size': None,
                            'modified': None,
                            'status': 'available'
                        }
                    ]
                }
                
    except Exception as e:
        logger.error(f"❌ VERBOSE: Error fetching Gemini models: {str(e)}")
        return {
            'status': 'error',
            'message': f'Error fetching models: {str(e)}',
            'fallback_models': [
                {
                    'name': 'gemini-2.0-flash',
                    'description': 'Latest multimodal model (fallback)',
                    'version': 'latest',
                    'size': None,
                    'modified': None,
                    'status': 'available'
                }
            ]
        } 